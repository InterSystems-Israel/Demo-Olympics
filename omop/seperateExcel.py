import iris
import ssl
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment

def main():
    connection_string = "k8s-6342c8de-a1773d26-900d2aa9ae-b0c6d6adee41ddb7.elb.us-east-1.amazonaws.com:443/USER"
    username = "SQLAdmin"
    password = "0m0pFHIR!"

    context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    context.verify_mode = ssl.CERT_REQUIRED
    context.check_hostname = False
    #context.load_verify_locations(r"C:\Users\nfrankel.ISCINTERNAL\OneDrive - InterSystems Corporation\Documents\OMOP Demo\certificateSQLaaS.pem")
    context.load_verify_locations(r"certificateSQLaaS.pem")
    connection = iris.connect(connection_string, username, password, sslcontext=context)
    print(connection)
    cursor = connection.cursor()

    # Get all tables in the schema
    schema = "OMOPCDM54"
    cursor.execute(f"SELECT table_name FROM information_schema.tables WHERE table_schema = '{schema}'")

    # Fetch all table names
    tables = cursor.fetchall()

    # Directory to save the Excel files
    output_directory = "omop_data_formatted_excel"
    os.makedirs(output_directory, exist_ok=True)

    # Create an Excel workbook
    workbook = Workbook()

    # Iterate through the tables and fetch data
    for (table_name,) in tables:
        try:
            print(f"Fetching data from table: {table_name}")
            # Query to fetch data
            query = f"SELECT TOP 60000 * FROM {schema}.{table_name}"
            cursor.execute(query)
            results = cursor.fetchall()

            # Create a sheet in the workbook for the table
            sheet = workbook.create_sheet(title=table_name)

            # Write column headers
            columns = [desc[0] for desc in cursor.description]
            sheet.append(columns)

            # Write data rows
            for row in results:
                sheet.append(row)

            # Apply formatting
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                for cell in column_cells:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            print(f"Data from table {table_name} written to the Excel sheet.")
        except Exception as e:
            print(f"Error fetching data from table {table_name}: {e}")

    # Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        std = workbook['Sheet']
        workbook.remove(std)

    # Save the workbook
    excel_file_path = os.path.join(output_directory, "omop_data2.xlsx")
    workbook.save(excel_file_path)
    print(f"Data written to {excel_file_path}")

    # Close the cursor and connection
    cursor.close()
    connection.close()

if __name__ == "__main__":
    main()
